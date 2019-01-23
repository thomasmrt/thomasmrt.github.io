"""
"""
import os
from flask_frozen import Freezer
from flask import Flask, render_template, redirect, request, session,  url_for, Markup
from openpyxl import load_workbook
from path import Path

def get_data(name):
    PATH = Path(r'static\data')
    wb = load_workbook(filename = PATH / 'data.xlsx')
    ws = wb[name]
    data = {'stages':[], 'main_title':None, 'sub_title':None, 'background': None}
    i = 2
    while ws.cell(row=i, column=1).value != None:
        if ws.cell(row=i, column=1).value == 'main_info':
            data['main_title'] = ws.cell(row=i, column=2).value
            data['sub_title'] = ws.cell(row=i, column=3).value
            data['background'] = ws.cell(row=i, column=4).value
            data['date'] = ws.cell(row=i, column=5).value
        elif ws.cell(row=i, column=1).value == 'section_start':            
            stage = {'title': None, 'buble': None,'content':[], 'info': None}
    

        elif ws.cell(row=i, column=1).value == 'section_end':
            data['stages'].append(stage)

        
        elif ws.cell(row=i, column=1).value == 'title':
            stage['title'] = ws.cell(row=i, column=3).value
            stage['buble'] = ws.cell(row=i, column=2).value

        elif ws.cell(row=i, column=1).value == 'paragraph':
            stage['content'].append({'type': 'text','value': ws.cell(row=i, column=2).value })
            
        elif ws.cell(row=i, column=1).value == 'image':
            image =  []
            k = 2
            while  ws.cell(row=i, column=k).value:
                image.append(ws.cell(row=i, column=k).value)
                k+=1
            
            stage['content'].append({'type': 'image','value': image, 'nb': len(image), 'row': int(12/len(image)) })

       
        elif ws.cell(row=i, column=1).value == 'info':
            deniv = ws.cell(row=i, column=4).value
            temps = ws.cell(row=i, column=5).value
            distance = ws.cell(row=i, column=3).value
            gpx = ws.cell(row=i, column=6).value
            stage['info'] = [distance, deniv, temps, gpx]
    
        
        i+=1
        print(i, ws.cell(row=i, column=1).value)

    return data

app = Flask(__name__)
app.secret_key = b'_you_will_never_guess'


@app.route('/')
@app.route('/index.html')
def home():
    data = get_data(name='Taillefer')
    print(data)
    return render_template('index.html', data=data)

if __name__ == "__main__":
#    data = get_data()
#    print(data)
#    app.run(debug=True)
    freezer = Freezer(app)
#    freezer.run()
    freezer.freeze()